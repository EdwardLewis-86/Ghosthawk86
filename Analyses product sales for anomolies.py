import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# File paths
input_path = r"C:\Users\edwardl\OneDrive - Motus Corporation\Finance\2. Finance\5. Actuaries\0. Products\RStudio_Python Inputs\product_sales.xlsb"
output_path = r"C:\Users\edwardl\OneDrive - Motus Corporation\Finance\2. Finance\5. Actuaries\0. Products\RStudio_Python Outputs\product_sales_with_stats.xlsx"

# Load data
df = pd.read_excel(input_path, sheet_name="Sheet1", engine="pyxlsb")
df.columns = df.iloc[0]
df = df[1:].reset_index(drop=True)

# Identify and convert year columns
year_columns = [col for col in df.columns if str(col).strip().isdigit() and 2012 <= int(col) <= 2025]
year_columns = [int(col) for col in year_columns]
df[year_columns] = df[year_columns].apply(pd.to_numeric, errors='coerce')

# Pre-2024 years
pre2024_years = [y for y in year_columns if y <= 2023]

# Statistical calculations
# Custom mean function excluding zero values
def mean_excluding_zeros(series):
    non_zero = series[series != 0]
    return non_zero.mean() if not non_zero.empty else 0

# Pre-2024 years
pre2024_years = [y for y in year_columns if y <= 2023]

# Custom mean excluding zeros
def mean_excluding_zeros(series):
    non_zero = series[series != 0]
    return non_zero.mean() if not non_zero.empty else 0

# Base metrics
df["total_sales_pre2024"] = df[pre2024_years].sum(axis=1)
df["nonzero_mean_sales_pre2024"] = df[pre2024_years].apply(mean_excluding_zeros, axis=1)

# Annualize 2025 (assumes Jan–Mar only available)
df["annualized_2025"] = df[2025] * 4

# Total & Mean
df["total_sales_2024_2025"] = df[2024] + df[2025]
df["mean_sales_2024_2025"] = (df[2024] + df["annualized_2025"]) / 2

# Growth metrics
df["%_growth_total_sales"] = (
    (df["total_sales_2024_2025"] - df["total_sales_pre2024"]) / df["total_sales_pre2024"]
) * 100

df["%_growth_mean_sales"] = (
    (df["mean_sales_2024_2025"] - df["nonzero_mean_sales_pre2024"]) / df["nonzero_mean_sales_pre2024"]
) * 100

# Volatility & z-scores
df["std_sales_pre2024"] = df[pre2024_years].std(axis=1)
df["max_sales_pre2024"] = df[pre2024_years].max(axis=1)
df["median_sales_pre2024"] = df[pre2024_years].median(axis=1)
df["cv_pre2024"] = df["std_sales_pre2024"] / df["nonzero_mean_sales_pre2024"]
df["z_2024"] = (df[2024] - df["nonzero_mean_sales_pre2024"]) / df["std_sales_pre2024"]
df["z_2025"] = (df["annualized_2025"] - df["nonzero_mean_sales_pre2024"]) / df["std_sales_pre2024"]

# Apply to each row
df["nonzero_mean_sales_pre2024"] = df[pre2024_years].apply(mean_excluding_zeros, axis=1)
df["std_sales_pre2024"] = df[pre2024_years].std(axis=1)
df["total_sales_pre2024"] = df[pre2024_years].sum(axis=1)
df["max_sales_pre2024"] = df[pre2024_years].max(axis=1)
df["median_sales_pre2024"] = df[pre2024_years].median(axis=1)
df["cv_pre2024"] = df["std_sales_pre2024"] / df["nonzero_mean_sales_pre2024"]

# Updated % change and z-score calculations
df["%_chg_2024_vs_mean"] = ((df[2024] - df["nonzero_mean_sales_pre2024"]) / df["nonzero_mean_sales_pre2024"]) * 100
df["%_chg_2025_vs_mean"] = ((df["annualized_2025"] - df["nonzero_mean_sales_pre2024"]) / df["nonzero_mean_sales_pre2024"]) * 100
df["z_2024"] = (df[2024] - df["nonzero_mean_sales_pre2024"]) / df["std_sales_pre2024"]
df["z_2025"] = (df[2025] - df["nonzero_mean_sales_pre2024"]) / df["std_sales_pre2024"]

# Build multi-criteria Anomaly_Flag
def generate_flags(row):
    flags = []

    # Always apply Dormant-To-Active logic
    if row["total_sales_pre2024"] == 0 and (row[2024] > 100 or row[2025] > 100):
        flags.append("Dormant-To-Active")

    # Only proceed with other flags if z-scores are valid
    z_2024_valid = pd.notna(row["z_2024"]) and not np.isinf(row["z_2024"])
    z_2025_valid = pd.notna(row["z_2025"]) and not np.isinf(row["z_2025"])

    if z_2024_valid or z_2025_valid:
        if (z_2024_valid and row["z_2024"] > 1.5) or (z_2025_valid and row["z_2025"] > 1.5):
            flags.append("Z-Score")
        if row["cv_pre2024"] > 1:
            flags.append("High CV")
        if row["%_chg_2024_vs_mean"] > 300 or row["%_chg_2025_vs_mean"] > 300:
            flags.append("Extreme Growth")

    return " + ".join(flags)

df["Anomaly_Flag"] = df.apply(generate_flags, axis=1)

output_columns = [
    "PRD_Name", "Product_Variant", "Product_Plan_Name", "RTF_TermPeriod",
    "Mechanical_Breakdown_Plan", "Total_Summary_Premium", "Grand Total"
] + year_columns + ["annualized_2025"] + [
    "total_sales_pre2024", "total_sales_2024_2025", "%_growth_total_sales",
    "nonzero_mean_sales_pre2024", "mean_sales_2024_2025", "%_growth_mean_sales",
    "std_sales_pre2024", "max_sales_pre2024", "median_sales_pre2024", "cv_pre2024",
    "%_chg_2024_vs_mean", "%_chg_2025_vs_mean", "z_2024", "z_2025", "Anomaly_Flag"
]


# Save to Excel
df_output = df[output_columns]
df_output.to_excel(output_path, index=False)

# Apply Excel formatting
wb = load_workbook(output_path)
ws = wb.active

# Bold header
for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save(output_path)

print(f"✅ Excel exported with multi-factor anomaly flags:\n{output_path}")

# === Step 1: Filter anomalies for summary tab ===
target_flags = [
    "Z-Score + High CV + Extreme Growth",
    "Z-Score + Extreme Growth",
    "High CV + Extreme Growth",
    "Z-Score + High CV",
    "Dormant-To-Active + Z-Score + Extreme Growth"
]

summary_df = df_output[df_output["Anomaly_Flag"].isin(target_flags)].copy()

# === Step 2: Add as second worksheet ===
from openpyxl import load_workbook

wb = load_workbook(output_path)
ws_summary = wb.create_sheet("Summary_Anomalies")

# Write DataFrame to worksheet
for r in dataframe_to_rows(summary_df, index=False, header=True):
    ws_summary.append(r)

# Bold header
for cell in ws_summary[1]:
    cell.font = Font(bold=True)


# Save final workbook
wb.save(output_path)
print(f"✅ Summary tab created with top-level anomalies:\n{output_path}")
